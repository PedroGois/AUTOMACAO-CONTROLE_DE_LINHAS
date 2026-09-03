"""Processa a base de telefonia e atualiza os dados consolidados do Dashboard."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


# ============================================================
# 1. CONFIGURAÇÃO DOS CAMINHOS E ARQUIVOS
# ============================================================
# Define onde o script está, onde o projeto está e quais arquivos
# serão usados para leitura e gravação dos dados do dashboard.
PASTA_SCRIPT = Path(__file__).resolve().parent
PASTA_PROJETO = PASTA_SCRIPT.parent
ARQUIVO_PADRAO = PASTA_PROJETO / "data" / "entrada" / "TELEFONIA.xlsx"
SAIDAS_DIR = PASTA_PROJETO / "data" / "saidas"
DASHBOARD_DIR = PASTA_PROJETO / "dashboard"
DADOS_DIR = DASHBOARD_DIR / "dados"
DADOS_JS_PADRAO = DADOS_DIR / "dados_dashboard.js"
DADOS_JSON_PADRAO = DADOS_DIR / "dados_dashboard.json"
HISTORICO_PATH = DADOS_DIR / "historico_dashboard.json"
STATUS_EXIBIDOS = {"ATIVA", "ESTOQUE", "DESLIGADO", "VERIFICAR"}
ABA_APARELHOS = "Celulares - Patrimônio"
ABA_PARCELAMENTOS = "Parcelamentos"
STATUS_PARCELAMENTO = {"PAGANDO", "PAGO"}


# ============================================================
# 2. FUNÇÕES DE LIMPEZA E PADRONIZAÇÃO
# ============================================================
# Estas funções deixam os textos iguais para facilitar comparação
# e evitar erros por causa de acentos, espaços ou letras minúsculas.
def normalizar(valor: object) -> str:
    texto = "" if valor is None else str(valor).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(letra for letra in texto if not unicodedata.combining(letra))
    return re.sub(r"\s+", " ", texto)


def texto(valor: object) -> str:
    # Remove espaços extras e deixa o texto limpo para gravar no JSON.
    return "" if valor is None else str(valor).strip()


def numero(valor: object) -> float:
    if isinstance(valor, (int, float)):
        return float(valor)
    if valor in (None, ""):
        return 0.0
    try:
        return float(str(valor).replace("R$", "").replace(".", "").replace(",", ".").strip())
    except (TypeError, ValueError):
        return 0.0


def obter_competencia(valor: object) -> str:
    """Converte datas da planilha para a chave mensal AAAA-MM."""
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m")
    bruto = texto(valor)
    if not bruto:
        return ""
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m"):
        try:
            return datetime.strptime(bruto[:10], formato).strftime("%Y-%m")
        except ValueError:
            continue
    encontrado = re.search(r"(20\d{2})[-/](0?[1-9]|1[0-2])", bruto)
    return f"{encontrado.group(1)}-{int(encontrado.group(2)):02d}" if encontrado else ""


def avancar_competencia(periodo: str) -> str:
    """Parcelamentos de um mês entram no fechamento de linhas do mês seguinte."""
    if not periodo:
        return ""
    ano, mes = (int(parte) for parte in periodo.split("-"))
    return f"{ano + 1}-01" if mes == 12 else f"{ano}-{mes + 1:02d}"


def meses_de_parcelamento(periodo: object, data_compra: object) -> int:
    """Calcula a parcela quando a fórmula do Excel ainda não foi recalculada."""
    if not isinstance(periodo, (datetime, date)) or not isinstance(data_compra, (datetime, date)):
        return 0
    return max(0, (periodo.year - data_compra.year) * 12 + periodo.month - data_compra.month + 1)


def valor_coluna(linha, colunas: dict[str, int], nome: str) -> object:
    indice = colunas.get(nome)
    return linha[indice - 1] if indice and indice <= len(linha) else None


def chave_cdc(valor: object) -> str:
    """Padroniza o código para relacionar a base à aba CDC."""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return texto(valor).upper()


def carregar_centros_custo(wb) -> dict[str, str]:
    """Lê a tabela de referência CDC sem depender do recálculo do Excel."""
    if "CDC" not in wb.sheetnames:
        return {}

    ws = wb["CDC"]
    colunas = indice_colunas(ws)
    col_codigo = colunas.get("CODCCUSTO")
    col_nome = colunas.get("NOME")
    if not col_codigo or not col_nome:
        return {}

    return {
        chave_cdc(linha[col_codigo - 1]): texto(linha[col_nome - 1])
        for linha in ws.iter_rows(min_row=2, values_only=True)
        if chave_cdc(linha[col_codigo - 1]) and texto(linha[col_nome - 1])
    }


def nome_centro_custo(centros_custo: dict[str, str], codigo: object, valor: object) -> str:
    """Usa o valor da planilha ou recompõe CDC quando a fórmula não tiver cache."""
    return texto(valor) or centros_custo.get(chave_cdc(codigo), "") or "SEM CENTRO DE CUSTO"


def indice_colunas(ws) -> dict[str, int]:
    """Devolve os índices das colunas pelo nome, ignorando acentos e espaços."""
    return {normalizar(celula.value): celula.column for celula in ws[1] if celula.value}


# ============================================================
# 3. VALIDAÇÃO DAS COLUNAS DA PLANILHA
# ============================================================
# Garante que a aba "Planos" tenha as colunas mínimas necessárias.
def localizar_colunas(ws) -> dict[str, int]:
    indices = indice_colunas(ws)
    obrigatorias = {"EMPRESA", "LINHA", "CHAPA/CPF", "NOME", "CPF", "COD CDC", "CDC", "STATUS", "DT FECHAMENTO"}
    ausentes = obrigatorias - indices.keys()
    if ausentes:
        raise ValueError("Colunas ausentes na aba Planos: " + ", ".join(sorted(ausentes)))
    return indices


# ============================================================
# 4. LEITURA E TRANSFORMAÇÃO DOS DADOS
# ============================================================
# Abre a planilha, lê somente os registros relevantes e monta uma lista
# com os dados prontos para o dashboard.
def carregar_dados(arquivo: Path) -> list[dict[str, object]]:
    wb = load_workbook(arquivo, read_only=True, data_only=True)
    try:
        if "Planos" not in wb.sheetnames:
            raise ValueError("A aba 'Planos' nao foi encontrada.")

        ws = wb["Planos"]
        colunas = localizar_colunas(ws)
        centros_custo = carregar_centros_custo(wb)

        # Procura a coluna de valor caso exista. Se não existir, usa zero.
        col_valor = colunas.get("VALOR") or {normalizar(c.value): c.column for c in ws[1] if c.value}.get("VALOR")

        registros = []
        for valores in ws.iter_rows(min_row=2, values_only=True):
            # Ignora linhas vazias.
            if not any(valor not in (None, "") for valor in valores):
                continue

            competencia = obter_competencia(valores[colunas["DT FECHAMENTO"] - 1])
            status = normalizar(valores[colunas["STATUS"] - 1])
            if not competencia or status not in STATUS_EXIBIDOS:
                continue

            empresa = normalizar(valores[colunas["EMPRESA"] - 1])

            # Converte o valor da linha para número, mesmo quando vem com moeda.
            val_num = 0.0
            if col_valor:
                val_raw = valores[col_valor - 1]
                if isinstance(val_raw, (int, float)):
                    val_num = float(val_raw)
                elif val_raw:
                    try:
                        limpo = str(val_raw).replace("R$", "").replace(".", "").replace(",", ".").strip()
                        val_num = float(limpo)
                    except ValueError:
                        val_num = 0.0

            registros.append({
                "competencia": competencia,
                "operadora": empresa if empresa in {"VIVO", "TIM"} else "OUTRAS",
                "linha": texto(valores[colunas["LINHA"] - 1]),
                "chapaCpf": texto(valores[colunas["CHAPA/CPF"] - 1]),
                "nome": texto(valores[colunas["NOME"] - 1]),
                "cpf": texto(valores[colunas["CPF"] - 1]),
                "codCdc": texto(valores[colunas["COD CDC"] - 1]) or "SEM CODIGO",
                "cdc": nome_centro_custo(
                    centros_custo,
                    valores[colunas["COD CDC"] - 1],
                    valores[colunas["CDC"] - 1],
                ),
                "status": status,
                "valor": val_num,
            })
        return registros
    finally:
        wb.close()


def carregar_aparelhos(arquivo: Path) -> list[dict[str, str]]:
    """Lê a aba patrimonial para controle visual; não altera a automação de linhas."""
    wb = load_workbook(arquivo, read_only=True, data_only=True)
    try:
        if ABA_APARELHOS not in wb.sheetnames:
            return []

        ws = wb[ABA_APARELHOS]
        colunas = indice_colunas(ws)
        obrigatorias = {"EMPRESA", "COD CDC", "CDC", "NUM DO PATRIMONIO"}
        ausentes = obrigatorias - colunas.keys()
        if ausentes:
            raise ValueError("Colunas ausentes na aba de aparelhos: " + ", ".join(sorted(ausentes)))

        def valor(linha, nome: str) -> str:
            indice = colunas.get(nome)
            return texto(linha[indice - 1]) if indice else ""

        aparelhos = []
        for linha in ws.iter_rows(min_row=2, values_only=True):
            if not any(v not in (None, "") for v in linha):
                continue
            status = normalizar(valor(linha, "STATUS")) or "VERIFICAR"
            if status not in STATUS_EXIBIDOS:
                status = "VERIFICAR"
            aparelhos.append({
                "modelo": valor(linha, "EMPRESA") or "SEM MODELO",
                "linha": valor(linha, "LINHA"),
                "codCdc": valor(linha, "COD CDC") or "SEM CODIGO",
                "cdc": valor(linha, "CDC") or "SEM CENTRO DE CUSTO",
                "chapa": valor(linha, "CHAPA"),
                "nome": valor(linha, "NOME"),
                "patrimonio": valor(linha, "NUM DO PATRIMONIO") or "SEM PATRIMONIO",
                "serie": valor(linha, "NUM SERIE"),
                "imei": valor(linha, "IMEI 1"),
                "status": status,
            })
        return aparelhos
    finally:
        wb.close()


def carregar_parcelamentos(arquivo: Path) -> list[dict[str, object]]:
    """Lê o fechamento mensal de aparelhos e preserva a competência de cada registro."""
    wb = load_workbook(arquivo, read_only=True, data_only=True)
    try:
        if ABA_PARCELAMENTOS not in wb.sheetnames:
            return []
        ws = wb[ABA_PARCELAMENTOS]
        colunas = indice_colunas(ws)
        obrigatorias = {
            "PERIODO", "NUM CONTA", "EMPRESA", "LINHA", "COD CDC", "CDC", "NOME",
            "NUM SERIE", "DATA DA COMPRA", "VALOR TOTAL", "VALOR MENSAL", "PARCELAMENTO",
            "NUM PARCELAS", "STATUS", "TERMO",
        }
        ausentes = obrigatorias - colunas.keys()
        if ausentes:
            raise ValueError("Colunas ausentes na aba Parcelamentos: " + ", ".join(sorted(ausentes)))

        registros = []
        for linha in ws.iter_rows(min_row=2, values_only=True):
            if not any(v not in (None, "") for v in linha):
                continue
            periodo_origem = obter_competencia(valor_coluna(linha, colunas, "PERIODO"))
            if not periodo_origem:
                continue
            periodo = valor_coluna(linha, colunas, "PERIODO")
            data_compra = valor_coluna(linha, colunas, "DATA DA COMPRA")
            valor_total = numero(valor_coluna(linha, colunas, "VALOR TOTAL"))
            num_parcelas = int(numero(valor_coluna(linha, colunas, "NUM PARCELAS")))
            valor_mensal = numero(valor_coluna(linha, colunas, "VALOR MENSAL"))
            parcela_atual = int(numero(valor_coluna(linha, colunas, "PARCELAMENTO")))
            status = normalizar(valor_coluna(linha, colunas, "STATUS"))

            # A aba Parcelamentos usa fórmulas. Quando o arquivo é salvo sem
            # recálculo pelo Excel, openpyxl recebe os resultados como vazios.
            if not valor_mensal and valor_total and num_parcelas:
                valor_mensal = valor_total / num_parcelas
            if not parcela_atual:
                parcela_atual = meses_de_parcelamento(periodo, data_compra)
            if not status and parcela_atual and num_parcelas:
                status = "PAGANDO" if parcela_atual <= num_parcelas else "PAGO"

            registros.append({
                "competencia": avancar_competencia(periodo_origem),
                "periodoOrigem": periodo_origem,
                "conta": texto(valor_coluna(linha, colunas, "NUM CONTA")),
                "empresa": normalizar(valor_coluna(linha, colunas, "EMPRESA")) or "VIVO",
                "linha": texto(valor_coluna(linha, colunas, "LINHA")),
                "codCdc": texto(valor_coluna(linha, colunas, "COD CDC")) or "SEM CODIGO",
                "cdc": texto(valor_coluna(linha, colunas, "CDC")) or "SEM CENTRO DE CUSTO",
                "chapa": texto(valor_coluna(linha, colunas, "CHAPA")),
                "nome": texto(valor_coluna(linha, colunas, "NOME")),
                "serie": texto(valor_coluna(linha, colunas, "NUM SERIE")),
                "dataCompra": texto(data_compra),
                "valorTotal": valor_total,
                "valorMensal": valor_mensal,
                "parcelaAtual": parcela_atual,
                "numParcelas": num_parcelas,
                "status": status,
                "termo": normalizar(valor_coluna(linha, colunas, "TERMO")),
                "atualizadoEm": texto(valor_coluna(linha, colunas, "DT ATUALIZACAO")),
            })
        return registros
    finally:
        wb.close()


def validar_dados(dados: list[dict[str, object]], parcelamentos: list[dict[str, object]]) -> list[dict[str, str]]:
    """Produz alertas acionáveis sem impedir a publicação do dashboard."""
    alertas = []
    competencias_linhas = {str(item["competencia"]) for item in dados}
    competencias_parcelas = {str(item["competencia"]) for item in parcelamentos}
    somente_linhas = sorted(competencias_linhas - competencias_parcelas)
    somente_parcelas = sorted(competencias_parcelas - competencias_linhas)
    if somente_linhas:
        alertas.append({"tipo": "aviso", "mensagem": "Competências somente em Planos: " + ", ".join(somente_linhas)})
    if somente_parcelas:
        alertas.append({"tipo": "aviso", "mensagem": "Competências somente em Parcelamentos: " + ", ".join(somente_parcelas)})

    chaves_linhas = Counter(f'{d["competencia"]}|{d["operadora"]}|{d["linha"]}' for d in dados)
    qtd_duplicadas = sum(1 for quantidade in chaves_linhas.values() if quantidade > 1)
    if qtd_duplicadas:
        alertas.append({"tipo": "erro", "mensagem": f"{qtd_duplicadas} linha(s) duplicada(s) dentro da mesma competência."})

    chaves_parcelas = Counter()
    falhas = Counter()
    for item in parcelamentos:
        identificador = item["serie"] or f'{item["empresa"]}|{item["linha"]}|{item["conta"]}'
        chaves_parcelas[f'{item["competencia"]}|{identificador}'] += 1
        if not item["dataCompra"]:
            falhas["data da compra"] += 1
        if not item["numParcelas"]:
            falhas["número de parcelas"] += 1
        if not item["valorMensal"]:
            falhas["valor mensal"] += 1
        if item["status"] not in STATUS_PARCELAMENTO:
            falhas["status"] += 1
        if item["termo"] not in {"", "SIM", "NAO"}:
            falhas["termo"] += 1
    qtd_duplicadas = sum(1 for quantidade in chaves_parcelas.values() if quantidade > 1)
    if qtd_duplicadas:
        alertas.append({"tipo": "erro", "mensagem": f"{qtd_duplicadas} parcelamento(s) duplicado(s) dentro da mesma competência."})
    for campo, quantidade in falhas.items():
        alertas.append({"tipo": "erro", "mensagem": f"{quantidade} parcelamento(s) com {campo} inválido ou vazio."})
    return alertas


# ============================================================
# 5. HISTÓRICO E COMPARATIVO
# ============================================================
# Cria o histórico de execuções e compara com o último registro salvo
# para mostrar a diferença de custo entre as medições.
def processar_historico(dados: list[dict[str, object]], gerado_em: datetime) -> dict[str, object]:
    """Salva a medição atual no histórico e retorna a comparação com a anterior."""
    HISTORICO_PATH.parent.mkdir(parents=True, exist_ok=True)

    # Soma os dados atuais por status para o dashboard e para o histórico.
    totais_atuais = {
        "data": gerado_em.strftime("%d/%m/%Y às %H:%M"),
        "total_linhas": len(dados),
        "custo_total": sum(float(d["valor"]) for d in dados),
        "ATIVA": {"qtd": 0, "val": 0.0},
        "ESTOQUE": {"qtd": 0, "val": 0.0},
        "DESLIGADO": {"qtd": 0, "val": 0.0},
        "VERIFICAR": {"qtd": 0, "val": 0.0},
    }
    for d in dados:
        st = str(d["status"])
        if st in totais_atuais:
            totais_atuais[st]["qtd"] += 1
            totais_atuais[st]["val"] += float(d["valor"])

    # Carrega o histórico salvo anteriormente.
    historico = []
    if HISTORICO_PATH.exists():
        try:
            historico = json.loads(HISTORICO_PATH.read_text(encoding="utf-8"))
            if not isinstance(historico, list):
                historico = []
        except Exception:
            historico = []

    anterior = historico[-1] if historico else None

    def calc_delta(atual_val: float, ant_val: float) -> dict[str, object]:
        diff = atual_val - ant_val
        pct = (diff / ant_val * 100) if ant_val > 0 else 0.0
        return {"diff": diff, "pct": round(pct, 1), "tem_anterior": ant_val > 0}

    comparativo = {
        "tem_anterior": anterior is not None,
        "data_anterior": anterior.get("data") if anterior else None,
        "custo_total": calc_delta(totais_atuais["custo_total"], anterior.get("custo_total", 0.0) if anterior else 0.0),
        "ATIVA": calc_delta(totais_atuais["ATIVA"]["val"], anterior.get("ATIVA", {}).get("val", 0.0) if anterior else 0.0),
        "ESTOQUE": calc_delta(totais_atuais["ESTOQUE"]["val"], anterior.get("ESTOQUE", {}).get("val", 0.0) if anterior else 0.0),
        "DESLIGADO": calc_delta(totais_atuais["DESLIGADO"]["val"], anterior.get("DESLIGADO", {}).get("val", 0.0) if anterior else 0.0),
        "VERIFICAR": calc_delta(totais_atuais["VERIFICAR"]["val"], anterior.get("VERIFICAR", {}).get("val", 0.0) if anterior else 0.0),
    }

    # Evita salvar o mesmo registro duas vezes durante a mesma execução.
    if not historico or historico[-1].get("data") != totais_atuais["data"]:
        historico.append(totais_atuais)
        if len(historico) > 30:
            historico = historico[-30:]
        HISTORICO_PATH.write_text(json.dumps(historico, ensure_ascii=False, indent=2), encoding="utf-8")

    return comparativo


# ============================================================
# 6. EXPORTAÇÃO DOS ARQUIVOS DO DASHBOARD
# ============================================================
# Gera os arquivos finais em JSON e JS para o dashboard consumir.
def exportar_dados(
    dados: list[dict[str, object]],
    aparelhos: list[dict[str, str]],
    parcelamentos: list[dict[str, object]],
    validacoes: list[dict[str, str]],
    gerado_em: datetime,
) -> None:
    """Exporta os dados em formato .js (para abrir direto com duplo clique) e .json."""
    DADOS_DIR.mkdir(parents=True, exist_ok=True)

    competencias_linhas = sorted({str(item["competencia"]) for item in dados})
    competencias_parcelamentos = sorted({str(item["competencia"]) for item in parcelamentos})
    payload = {
        "gerado_em": gerado_em.strftime("%d/%m/%Y às %H:%M"),
        "total_registros": len(dados),
        "competencias": sorted(set(competencias_linhas) | set(competencias_parcelamentos)),
        "competencias_linhas": competencias_linhas,
        "competencias_parcelamentos": competencias_parcelamentos,
        "validacoes": validacoes,
        "dados": dados,
        "aparelhos": aparelhos,
        "parcelamentos": parcelamentos,
    }

    json_str = json.dumps(payload, ensure_ascii=False, indent=2)

    # Salva o arquivo principal em JSON.
    DADOS_JSON_PADRAO.write_text(json_str, encoding="utf-8")

    # Salva o arquivo JS para abrir a página sem problemas de CORS.
    js_content = f"window.DADOS_TELEFONIA = {json_str};\n"
    DADOS_JS_PADRAO.write_text(js_content, encoding="utf-8")


# ============================================================
# 7. EXECUÇÃO PRINCIPAL
# ============================================================
# Ponto de entrada do script. Lê a planilha, calcula os totais,
# salva os dados e imprime uma mensagem final no terminal.
def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--telefonia", type=Path, default=ARQUIVO_PADRAO)
    args = parser.parse_args()

    if not args.telefonia.exists():
        raise FileNotFoundError(f"Planilha de telefonia nao encontrada: {args.telefonia}")

    momento_atual = datetime.now()
    dados = carregar_dados(args.telefonia)
    aparelhos = carregar_aparelhos(args.telefonia)
    parcelamentos = carregar_parcelamentos(args.telefonia)
    validacoes = validar_dados(dados, parcelamentos)
    exportar_dados(dados, aparelhos, parcelamentos, validacoes, momento_atual)

    print("==================================================")
    print("Base de dados do Dashboard atualizada com sucesso!")
    # Caminhos relativos evitam falha de codificação em consoles Windows
    # quando a pasta do usuário contém caracteres especiais.
    print(f" • Arquivo JS:   {DADOS_JS_PADRAO.relative_to(PASTA_PROJETO)}")
    print(f" • Arquivo JSON: {DADOS_JSON_PADRAO.relative_to(PASTA_PROJETO)}")
    print(f" • Linhas carregadas: {len(dados)}")
    print(f" • Aparelhos carregados: {len(aparelhos)}")
    print(f" • Parcelamentos carregados: {len(parcelamentos)}")
    print(f" • Alertas de qualidade: {len(validacoes)}")
    print("==================================================")


if __name__ == "__main__":
    main()
