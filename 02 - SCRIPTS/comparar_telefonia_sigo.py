"""Compara a aba Planos com a base do SIGO de forma direta.

Regras de cruzamento (em ordem de prioridade):
  1. Chapa/Matricula (externalid) -- identificador mais estavel e unico.
  2. CPF              -- identificador nacional; usado quando a chapa nao bate.
  3. Nome normalizado -- fallback; sujeito a variacoes de grafia.

Linhas cuja Chapa/CPF contenha FROTA, FAMILIA ou FORA SIGO nao participam
da comparacao normal (comportamento preservado).
"""

from __future__ import annotations

import argparse
import re
import shutil
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PASTA_SCRIPT = Path(__file__).resolve().parent
PASTA_PROJETO = PASTA_SCRIPT.parent
PASTA_DADOS = PASTA_PROJETO / "01 - DADOS"
PASTA_SAIDAS = PASTA_PROJETO / "04 - SAIDAS"
SIGO_PADRAO = PASTA_SAIDAS / "BASE_SIGO.xlsx"
BACKUP_PADRAO = PASTA_SAIDAS / "BACKUPS"
RESULTADO_PADRAO = PASTA_SAIDAS / "resultado_comparacao.txt"
ABA_PLANOS = "Planos"
PARTICULAS_NOME = {"DA", "DAS", "DE", "DO", "DOS", "E"}


def localizar_telefonia() -> Path:
    teste = PASTA_DADOS / "TELEFONIA-TESTE.xlsx"
    return teste if teste.exists() else PASTA_DADOS / "TELEFONIA.xlsx"


def normalizar_nome(valor: object) -> str:
    """Remove acentos, espacos extras e particulas como de, da, dos e das."""
    texto = "" if valor is None else str(valor)
    texto = unicodedata.normalize("NFKD", texto.upper())
    texto = "".join(letra for letra in texto if not unicodedata.combining(letra))
    texto = re.sub(r"[^A-Z0-9 ]+", " ", texto)
    palavras = [palavra for palavra in texto.split() if palavra not in PARTICULAS_NOME]
    return " ".join(palavras)


def normalizar_chapa(valor: object) -> str:
    texto = "" if valor is None else str(valor).strip()
    if isinstance(valor, float) and valor.is_integer():
        texto = str(int(valor))
    return re.sub(r"[^A-Z0-9]", "", texto.upper())


def eh_frota(valor: object) -> bool:
    return "FROTA" in normalizar_nome(valor)


def eh_familia(valor: object) -> bool:
    return "FAMILIA" in normalizar_nome(valor)


def eh_fora_sigo(valor: object) -> bool:
    return "FORA SIGO" in normalizar_nome(valor)


def booleano(valor: object) -> bool | None:
    if isinstance(valor, bool):
        return valor
    texto = normalizar_nome(valor)
    if texto in {"TRUE", "VERDADEIRO", "SIM", "S", "1", "ATIVO"}:
        return True
    if texto in {"FALSE", "FALSO", "NAO", "N", "0", "INATIVO"}:
        return False
    return None


def cabecalhos(ws) -> dict[str, int]:
    return {
        normalizar_nome(celula.value): celula.column
        for celula in ws[1]
        if celula.value is not None
    }


def coluna(cab: dict[str, int], opcoes: list[str], descricao: str) -> int:
    for opcao in opcoes:
        encontrada = cab.get(normalizar_nome(opcao))
        if encontrada:
            return encontrada
    raise ValueError(f"Coluna obrigatoria nao encontrada: {descricao}")


def registro_unico(registros: list[dict]) -> dict | None:
    return registros[0] if len(registros) == 1 else None


def limpar_filtros(ws) -> None:
    """Remove criterios de filtro e reexibe todas as linhas antes da comparacao."""
    if ws.auto_filter:
        ws.auto_filter.filterColumn = []
        ws.auto_filter.sortState = None
    for linha in range(2, ws.max_row + 1):
        ws.row_dimensions[linha].hidden = False


def carregar_sigo(
    caminho: Path,
) -> tuple[dict[str, list[dict]], dict[str, list[dict]], dict[str, list[dict]]]:
    """Carrega a base SIGO e devolve tres indices: por chapa, por CPF e por nome."""
    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        cab = cabecalhos(ws)
        col_nome  = coluna(cab, ["nome"], "Nome")
        col_chapa = coluna(cab, ["externalid", "external id", "chapa"], "External ID")
        col_cpf   = coluna(cab, ["cpf", "cpf/cnpj"], "CPF")
        col_ativo = coluna(cab, ["isactive", "ativo"], "isactive")
        por_nome:  dict[str, list[dict]] = defaultdict(list)
        por_chapa: dict[str, list[dict]] = defaultdict(list)
        por_cpf:   dict[str, list[dict]] = defaultdict(list)
        for valores in ws.iter_rows(min_row=2, values_only=True):
            registro = {
                "nome":     str(valores[col_nome - 1] or "").strip(),
                "cpf":      re.sub(r"\D+", "", str(valores[col_cpf - 1] or "")),
                "isactive": booleano(valores[col_ativo - 1]),
            }
            nome  = normalizar_nome(valores[col_nome - 1])
            chapa = normalizar_chapa(valores[col_chapa - 1])
            if nome:
                por_nome[nome].append(registro)
            if chapa:
                por_chapa[chapa].append(registro)
            # Indexa tambem pelo CPF do SIGO (somente digitos, sem zeros a esquerda)
            cpf_sigo = re.sub(r"\D+", "", str(valores[col_cpf - 1] or "")).lstrip("0")
            if cpf_sigo:
                por_cpf[cpf_sigo].append(registro)
        return por_chapa, por_cpf, por_nome
    finally:
        wb.close()


def comparar(telefonia: Path, sigo: Path, backup_dir: Path, resultado: Path) -> None:
    # Desempacota os tres indices na nova ordem de prioridade: chapa, cpf, nome
    por_chapa, por_cpf, por_nome = carregar_sigo(sigo)
    wb = load_workbook(telefonia)
    try:
        if ABA_PLANOS not in wb.sheetnames:
            raise ValueError("A aba 'Planos' nao foi encontrada.")
        ws = wb[ABA_PLANOS]
        limpar_filtros(ws)
        cab = cabecalhos(ws)
        col_nome   = coluna(cab, ["nome"], "Nome")
        col_chapa  = coluna(cab, ["chapa/cpf", "chapa"], "Chapa/CPF")
        col_cpf    = coluna(cab, ["cpf"], "CPF")
        col_status = coluna(cab, ["status"], "Status")
        # Coluna Valor e opcional -- usada apenas para somar custo dos desligados
        col_valor  = cab.get(normalizar_nome("Valor"))
        col_empresa = cab.get(normalizar_nome("Empresa")) # <-- ADICIONE ESTA LINHA

        total_verificado = 0
        encontrados_chapa = encontrados_cpf = encontrados_nome = 0
        desligados = verificar = frotas = ambiguos = 0
        totais_operadora = defaultdict(lambda: defaultdict(float)) # <-- ADICIONE ESTA LINHA
        custo_desligados = 0.0  # soma dos valores mensais das linhas DESLIGADO

        def obter_valor(num_linha: int) -> float:
            if not col_valor:
                return 0.0
            val = ws.cell(num_linha, col_valor).value
            if isinstance(val, (int, float)):
                return float(val)
            try:
                # Trata casos de texto como "R$ 49,90" ou "49.90"
                texto_limpo = str(val or "0").replace("R$", "").replace(".", "").replace(",", ".").strip()
                return float(texto_limpo)
            except ValueError:
                return 0.0
            
        for linha in range(2, ws.max_row + 1):
            if not any(
                ws.cell(linha, c).value not in (None, "")
                for c in range(1, ws.max_column + 1)
            ):
                continue
            total_verificado += 1
            
            # Pega operadora e valor da linha atual
            empresa_raw = ws.cell(linha, col_empresa).value if col_empresa else "OUTRAS"
            operadora = str(empresa_raw or "OUTRAS").strip().upper()
            val_linha = obter_valor(linha)
            chapa_original = ws.cell(linha, col_chapa).value
            
            # Trata ESTOQUE
            if normalizar_nome(ws.cell(linha, col_status).value) == "ESTOQUE":
                totais_operadora[operadora]["ESTOQUE"] += val_linha
                continue
                
            # Trata FAMILIA
            if eh_familia(chapa_original):
                totais_operadora[operadora]["FAMILIA"] += val_linha
                continue
                
            # Trata FORA SIGO
            if eh_fora_sigo(chapa_original):
                ws.cell(linha, col_status).value = "ATIVA"
                totais_operadora[operadora]["ATIVA"] += val_linha
                continue
                
            # Trata FROTA
            if eh_frota(chapa_original):
                frotas += 1
                totais_operadora[operadora]["FROTA"] += val_linha
                continue
            # --- BUSCA NO SIGO (CHAPA -> CPF -> NOME) ---
            registro = registro_unico(por_chapa.get(normalizar_chapa(chapa_original), []))
            if registro:
                encontrados_chapa += 1
            else:
                cpf_tel = re.sub(r"\D+", "", str(ws.cell(linha, col_cpf).value or "")).lstrip("0")
                if cpf_tel:
                    registro = registro_unico(por_cpf.get(cpf_tel, []))
                    if registro:
                        encontrados_cpf += 1
            if registro is None:
                candidatos = por_nome.get(normalizar_nome(ws.cell(linha, col_nome).value), [])
                registro = registro_unico(candidatos)
                if registro:
                    encontrados_nome += 1
                elif len(candidatos) > 1:
                    ambiguos += 1
            # --- DEFINIÇÃO DO STATUS FINAL E SOMA ---
            if registro is None:
                ws.cell(linha, col_status).value = "VERIFICAR"
                verificar += 1
                totais_operadora[operadora]["VERIFICAR"] += val_linha
                continue
            if registro.get("nome"):
                ws.cell(linha, col_nome).value = registro["nome"]
            if registro["isactive"] is True:
                ws.cell(linha, col_cpf).value = registro["cpf"]
                status_atual = normalizar_nome(ws.cell(linha, col_status).value) or "ATIVA"
                totais_operadora[operadora][status_atual] += val_linha
            elif registro["isactive"] is False:
                ws.cell(linha, col_status).value = "DESLIGADO"
                desligados += 1
                custo_desligados += val_linha # <-- ADICIONE ESTA LINHA
                totais_operadora[operadora]["DESLIGADO"] += val_linha
            else:
                ws.cell(linha, col_status).value = "VERIFICAR"
                verificar += 1
                totais_operadora[operadora]["VERIFICAR"] += val_linha
            

        backup_dir.mkdir(parents=True, exist_ok=True)
        momento = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup  = backup_dir / f"{telefonia.stem}_antes_da_comparacao_{momento}{telefonia.suffix}"
        shutil.copy2(telefonia, backup)
        wb.save(telefonia)
    finally:
        wb.close()

    # resultado_comparacao.txt -- o menu le TOTAL_VERIFICADO (nao remover essa chave)
    resultado.write_text(
        f"TOTAL_VERIFICADO={total_verificado}\n"
        f"ENCONTRADOS_CHAPA={encontrados_chapa}\n"
        f"ENCONTRADOS_CPF={encontrados_cpf}\n"
        f"ENCONTRADOS_NOME={encontrados_nome}\n"
        f"DESLIGADOS={desligados}\n"
        f"VERIFICAR={verificar}\n"
        f"FROTAS={frotas}\n"
        f"AMBIGUOS={ambiguos}\n"
        f"CUSTO_DESLIGADOS={custo_desligados:.2f}\n",
        encoding="utf-8",
    )
    print(f"Planilha atualizada: {telefonia.name}")
    print(f"Total verificado (inclui FROTA): {total_verificado}")
    print(f"Encontrados por chapa: {encontrados_chapa}")        # 1a prioridade
    print(f"Encontrados por CPF:   {encontrados_cpf}")          # 2a prioridade
    print(f"Encontrados por nome:  {encontrados_nome}")         # 3a prioridade (fallback)
    print(f"Ambiguos (homonimos):  {ambiguos}")
    print(f"Desligados: {desligados}")
    if col_valor and custo_desligados:
        custo_fmt = f"{custo_desligados:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        print(f"Custo mensal desligados: R$ {custo_fmt}")
    print(f"Para verificar: {verificar}")
    print(f"Frota ignorada: {frotas}")
    print(f"Backup criado: {backup.name}")

    def formatar_moeda(valor: float) -> str:
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    print("\n" + "=" * 55)
    print(f"RESUMO FINANCEIRO POR OPERADORA E STATUS")
    print("=" * 55)
    for op, status_map in sorted(totais_operadora.items()):
        total_op = sum(status_map.values())
        print(f"\n[{op}] - Custo Total: {formatar_moeda(total_op)}")
        for st in ["ATIVA", "ESTOQUE", "DESLIGADO", "VERIFICAR", "FROTA", "FAMILIA"]:
            val = status_map.get(st, 0.0)
            if val > 0:
                print(f"   • {st:<12}: {formatar_moeda(val)}")
    print("=" * 55 + "\n")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--telefonia", type=Path, default=localizar_telefonia())
    parser.add_argument("--sigo", type=Path, default=SIGO_PADRAO)
    parser.add_argument("--backup", type=Path, default=BACKUP_PADRAO)
    parser.add_argument("--resultado", type=Path, default=RESULTADO_PADRAO)
    args = parser.parse_args()
    if not args.telefonia.exists():
        raise FileNotFoundError(f"Planilha de telefonia nao encontrada: {args.telefonia}")
    if not args.sigo.exists():
        raise FileNotFoundError(f"Base SIGO nao encontrada: {args.sigo}")
    comparar(args.telefonia, args.sigo, args.backup, args.resultado)


if __name__ == "__main__":
    main()
