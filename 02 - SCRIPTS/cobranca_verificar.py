"""Gera conferências por centro de custo e rascunhos para linhas com Status=VERIFICAR.

O projeto não envia mensagens. No modo outlook, apenas salva mensagens na pasta
Rascunhos do perfil local; no modo eml, cria arquivos de rascunho para revisão.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

STATUS_ALVO = "VERIFICAR"
CABECALHO = PatternFill("solid", fgColor="008C95")


def texto(valor: object) -> str:
    return "" if valor is None else str(valor).strip()


def chave(valor: object) -> str:
    return re.sub(r"\s+", " ", texto(valor)).casefold()


def seguro(valor: object, limite: int = 60) -> str:
    nome = re.sub(r"[^A-Za-z0-9]+", "_", texto(valor)).strip("_")
    return (nome or "SEM_CENTRO_DE_CUSTO")[:limite]


def ler_planos(arquivo: Path, empresa: str | None = None) -> tuple[list[str], list[dict[str, object]]]:
    wb = load_workbook(arquivo, read_only=True, data_only=True)
    try:
        if "Planos" not in wb.sheetnames:
            raise ValueError("A aba 'Planos' não foi encontrada.")
        ws = wb["Planos"]
        linhas = ws.iter_rows(values_only=True)
        cabecalho = [texto(valor) for valor in next(linhas)]
        while cabecalho and not cabecalho[-1]:
            cabecalho.pop()
        indice = {chave(nome): posicao for posicao, nome in enumerate(cabecalho)}
        obrigatorias = {"cod cdc", "cdc", "status"}
        if empresa:
            obrigatorias.add("empresa")
        ausentes = obrigatorias - indice.keys()
        if ausentes:
            raise ValueError(f"Colunas ausentes: {', '.join(sorted(ausentes))}")
        selecionadas = []
        for valores in linhas:
            registro = {nome: valores[posicao] if posicao < len(valores) else None
                        for nome, posicao in indice.items()}
            mesma_empresa = not empresa or chave(registro["empresa"]) == chave(empresa)
            if chave(registro["status"]) == chave(STATUS_ALVO) and mesma_empresa:
                selecionadas.append(registro)
        return cabecalho, selecionadas
    finally:
        wb.close()


def ler_contatos(arquivo: Path) -> dict[str, dict[str, str]]:
    wb = load_workbook(arquivo, read_only=True, data_only=True)
    try:
        ws = wb.active
        linhas = ws.iter_rows(values_only=True)
        cabecalho = [texto(valor) for valor in next(linhas)]
        indice = {chave(nome): posicao for posicao, nome in enumerate(cabecalho)}
        for coluna in ("cod cdc", "responsavel", "email"):
            if coluna not in indice:
                raise ValueError(f"A planilha de contatos não possui a coluna '{coluna}'.")
        contatos = {}
        for valores in linhas:
            codigo = texto(valores[indice["cod cdc"]])
            if codigo:
                contatos[codigo] = {
                    "responsavel": texto(valores[indice["responsavel"]]),
                    "email": texto(valores[indice["email"]]),
                }
        return contatos
    finally:
        wb.close()


def gerar_planilha(destino: Path, cabecalho: list[str], registros: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Verificar"
    ws.freeze_panes = "A2"
    ws.append(cabecalho)
    for celula in ws[1]:
        celula.fill = CABECALHO
        celula.font = Font(color="FFFFFF", bold=True)
        celula.alignment = Alignment(horizontal="center")
    for registro in registros:
        ws.append([registro.get(chave(coluna)) for coluna in cabecalho])
    for coluna in ws.columns:
        letra = coluna[0].column_letter
        maior = max(len(texto(celula.value)) for celula in coluna)
        ws.column_dimensions[letra].width = min(max(maior + 2, 12), 42)
    ws.auto_filter.ref = ws.dimensions
    tabela = Table(displayName="LinhasVerificar", ref=ws.dimensions)
    tabela.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tabela)
    wb.save(destino)
    wb.close()


def eh_suporte(responsavel: str, email: str) -> bool:
    """Verifica se o destinatário é do Suporte/TI interna."""
    texto_busca = f"{responsavel} {email}".casefold()
    return "suporte" in texto_busca
def corpo_email_suporte(centros: list[tuple[str, str, int]], prazo: str) -> str:
    """Texto exclusivo para quando a conferência for do próprio Suporte / TI."""
    total = sum(quantidade for _, _, quantidade in centros)
    itens = "\n".join(
        f"- CC {codigo} — {centro}: {quantidade} linha(s)"
        for codigo, centro, quantidade in centros
    )
    return f"""Olá, Suporte / TI.
Segue em anexo a relação de linhas com status VERIFICAR atribuídas ao setor de Suporte/TI para conferência interna.
Centros de Custo incluídos:
{itens}
Total: {total} linha(s) para conferência direta.
Ações necessárias até {prazo}:
1. Verificar no SIGO / chamados quem está com cada linha ou se estão em estoque/reserva técnica.
2. Atualizar a planilha principal (TELEFONIA.xlsx) com o status e responsável definitivos.
Atenciosamente,
Controle de Telefonia & Ativos
"""
def corpo_email_gestor(responsavel: str, centros: list[tuple[str, str, int]], prazo: str) -> str:
    """Texto padrão enviado para os gestores e responsáveis de CDC."""
    saudacao = responsavel or "pessoal"
    total = sum(quantidade for _, _, quantidade in centros)
    if len(centros) == 1:
        codigo, centro, quantidade = centros[0]
        escopo = (
            f"Estamos realizando a conferência das linhas corporativas vinculadas ao "
            f"centro de custo {centro} (CC {codigo}).\n"
            f"Segue anexa a relação de {quantidade} linha(s) marcada(s) como VERIFICAR "
            "para validação."
        )
    else:
        itens = "\n".join(
            f"- CC {codigo} — {centro}: {quantidade} linha(s)"
            for codigo, centro, quantidade in centros
        )
        escopo = (
            "Estamos realizando a conferência das linhas corporativas vinculadas aos "
            "centros de custo sob sua responsabilidade. Seguem anexas as relações para validação:\n\n"
            f"{itens}\n\nTotal: {total} linha(s) marcada(s) como VERIFICAR."
        )
    return f"""Olá, {saudacao}.
{escopo}
Por gentileza, informe o responsável correto por cada chip ou eventual ajuste necessário até o dia {prazo}.
Reforçamos que toda alteração de responsabilidade de chips ou linhas corporativas deve ser comunicada à TI, para manter o inventário atualizado e garantir a rastreabilidade dos ativos.
Ficamos à disposição.
Atenciosamente,
Equipe de TI
"""
def corpo_email(responsavel: str, email: str, centros: list[tuple[str, str, int]], prazo: str) -> str:
    """Seleciona o modelo correto de e-mail (Suporte vs Gestor)."""
    if eh_suporte(responsavel, email):
        return corpo_email_suporte(centros, prazo)
    return corpo_email_gestor(responsavel, centros, prazo)


def criar_eml(destino: Path, email: str, assunto: str, corpo: str,
               anexos: list[Path]) -> None:
    mensagem = EmailMessage()
    mensagem["To"] = email
    mensagem["Subject"] = assunto
    mensagem.set_content(corpo)
    for anexo in anexos:
        mensagem.add_attachment(anexo.read_bytes(), maintype="application",
                                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                filename=anexo.name)
    destino.write_bytes(mensagem.as_bytes())


def criar_outlook(email: str, assunto: str, corpo: str, anexos: list[Path]) -> None:
    try:
        import win32com.client  # type: ignore
    except ImportError as erro:
        raise RuntimeError("pywin32 não está instalado; use --modo eml ou instale pywin32.") from erro
    outlook = win32com.client.Dispatch("Outlook.Application")
    mensagem = outlook.CreateItem(0)
    mensagem.To = email
    mensagem.Subject = assunto
    mensagem.Body = corpo
    for anexo in anexos:
        mensagem.Attachments.Add(str(anexo.resolve()))
    mensagem.Save()


def executar(telefonia: Path, contatos_arquivo: Path, saida: Path, modo: str,
             prazo: str, empresa: str | None = None) -> None:
    cabecalho, linhas = ler_planos(telefonia, empresa)
    contatos = ler_contatos(contatos_arquivo)
    grupos: dict[str, list[dict[str, object]]] = defaultdict(list)
    nomes: dict[str, list[str]] = defaultdict(list)
    for linha in linhas:
        codigo = texto(linha["cod cdc"]) or "SEM_COD_CDC"
        grupos[codigo].append(linha)
        nomes[codigo].append(texto(linha["cdc"]) or "SEM CENTRO DE CUSTO")

    anexos = saida / "anexos"
    rascunhos = saida / "rascunhos_eml"
    anexos.mkdir(parents=True, exist_ok=True)
    if modo == "eml":
        rascunhos.mkdir(parents=True, exist_ok=True)
    for antigo in anexos.glob("VERIFICAR_*.xlsx"):
        antigo.unlink()
    if modo == "eml":
        for antigo in rascunhos.glob("*.eml"):
            antigo.unlink()
    relatorio = []
    envios: dict[str, dict[str, object]] = {}
    for codigo, registros in sorted(grupos.items()):
        centros = Counter(nomes[codigo])
        centro = centros.most_common(1)[0][0]
        base = f"VERIFICAR_{seguro(codigo, 20)}_{seguro(centro)}"
        anexo = anexos / f"{base}.xlsx"
        gerar_planilha(anexo, cabecalho, registros)
        contato = contatos.get(codigo, {})
        email = contato.get("email", "")
        pendencias = []
        if len(centros) > 1:
            pendencias.append("centro de custo com nomes divergentes")
        if not email:
            pendencias.append("sem e-mail cadastrado")
        if pendencias:
            relatorio.append([
                codigo, centro, contato.get("responsavel", ""), email,
                len(registros), "PENDENTE", "; ".join(pendencias), anexo.name,
            ])
            continue
        chave_email = chave(email)
        envio = envios.setdefault(chave_email, {
            "email": email,
            "responsavel": contato.get("responsavel", ""),
            "centros": [],
            "anexos": [],
        })
        envio["centros"].append((codigo, centro, len(registros)))
        envio["anexos"].append(anexo)

    for chave_email, envio in sorted(envios.items()):
        email = str(envio["email"])
        responsavel = str(envio["responsavel"])
        centros = list(envio["centros"])
        anexos_envio = list(envio["anexos"])
        codigos = ", ".join(codigo for codigo, _, _ in centros)
        if eh_suporte(responsavel, email):
            assunto = f"[CONFERÊNCIA INTERNA SUPORTE] Linhas corporativas — CC {codigos}"
        else:
            assunto = f"Conferência de linhas corporativas — CC {codigos}"
            
        corpo = corpo_email(responsavel, email, centros, prazo)
        base_email = f"COBRANCA_{seguro(responsavel or chave_email)}"
        if modo == "outlook":
            criar_outlook(email, assunto, corpo, anexos_envio)
            destino_rascunho = "Rascunhos do Outlook"
        else:
            arquivo_eml = rascunhos / f"{base_email}.eml"
            criar_eml(arquivo_eml, email, assunto, corpo, anexos_envio)
            destino_rascunho = arquivo_eml.name
        for (codigo, centro, quantidade), anexo in zip(centros, anexos_envio):
            relatorio.append([
                codigo, centro, responsavel, email, quantidade,
                "CRIADO", destino_rascunho, anexo.name,
            ])

    with (saida / "relatorio_execucao.csv").open("w", encoding="utf-8-sig", newline="") as arquivo:
        escritor = csv.writer(arquivo, delimiter=";")
        escritor.writerow([
            "Cod CDC", "CDC", "Responsável", "E-mail", "Linhas",
            "Status", "Resultado/Pendência", "Anexo",
        ])
        escritor.writerows(relatorio)
    print(f"Operadora: {empresa or 'TODAS'}")
    print(f"Linhas VERIFICAR: {len(linhas)}")
    print(f"Centros processados: {len(grupos)}")
    print(f"Rascunhos criados: {len(envios)}")
    print(f"Pendências: {sum(item[5] == 'PENDENTE' for item in relatorio)}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--telefonia", type=Path, required=True)
    parser.add_argument("--contatos", type=Path, required=True)
    parser.add_argument("--saida", type=Path, required=True)
    parser.add_argument("--modo", choices=("eml", "outlook"), default="eml")
    parser.add_argument("--prazo", required=True,
                        help="Prazo informado no corpo dos e-mails (DD/MM/AAAA).")
    parser.add_argument("--empresa", choices=("VIVO", "TIM"),
                        help="Gera somente os rascunhos da operadora informada. Não depende de filtro no Excel.")
    args = parser.parse_args()
    try:
        datetime.strptime(args.prazo, "%d/%m/%Y")
    except ValueError as erro:
        raise SystemExit("O prazo deve estar no formato DD/MM/AAAA.") from erro
    executar(args.telefonia, args.contatos, args.saida, args.modo, args.prazo, args.empresa)


if __name__ == "__main__":
    main()
