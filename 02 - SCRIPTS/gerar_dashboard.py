"""Processa a base de telefonia e atualiza os dados consolidados do Dashboard."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PASTA_SCRIPT = Path(__file__).resolve().parent
PASTA_PROJETO = PASTA_SCRIPT.parent
ARQUIVO_PADRAO = PASTA_PROJETO / "01 - DADOS" / "TELEFONIA.xlsx"
SAIDAS_DIR = PASTA_PROJETO / "04 - SAIDAS"
DADOS_JS_PADRAO = SAIDAS_DIR / "dados_dashboard.js"
DADOS_JSON_PADRAO = SAIDAS_DIR / "dados_dashboard.json"
HISTORICO_PATH = SAIDAS_DIR / "historico_dashboard.json"
STATUS_EXIBIDOS = {"ATIVA", "ESTOQUE", "DESLIGADO", "VERIFICAR"}


def normalizar(valor: object) -> str:
    texto = "" if valor is None else str(valor).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(letra for letra in texto if not unicodedata.combining(letra))
    return re.sub(r"\s+", " ", texto)


def texto(valor: object) -> str:
    return "" if valor is None else str(valor).strip()


def localizar_colunas(ws) -> dict[str, int]:
    indices = {normalizar(celula.value): celula.column for celula in ws[1] if celula.value}
    obrigatorias = {"EMPRESA", "LINHA", "CHAPA/CPF", "NOME", "CPF", "COD CDC", "CDC", "STATUS"}
    ausentes = obrigatorias - indices.keys()
    if ausentes:
        raise ValueError("Colunas ausentes na aba Planos: " + ", ".join(sorted(ausentes)))
    return indices


def carregar_dados(arquivo: Path) -> list[dict[str, object]]:
    wb = load_workbook(arquivo, read_only=True, data_only=True)
    try:
        if "Planos" not in wb.sheetnames:
            raise ValueError("A aba 'Planos' nao foi encontrada.")
        ws = wb["Planos"]
        colunas = localizar_colunas(ws)
        col_valor = colunas.get("VALOR") or {normalizar(c.value): c.column for c in ws[1] if c.value}.get("VALOR")
        
        registros = []
        for valores in ws.iter_rows(min_row=2, values_only=True):
            if not any(valor not in (None, "") for valor in valores):
                continue
            status = normalizar(valores[colunas["STATUS"] - 1])
            if status not in STATUS_EXIBIDOS:
                continue
            empresa = normalizar(valores[colunas["EMPRESA"] - 1])
            
            val_num = 0.0
            if col_valor:
                val_raw = valores[col_valor - 1]
                if isinstance(val_raw, (int, float)):
                    val_num = float(val_raw)
                elif val_raw:
                    try:
                        limpo = str(val_raw).replace("R$", "").replace(".", "").replace(",", ".").strip()
                        val_num = float(limpo)
                    except ValueError:
                        val_num = 0.0

            registros.append({
                "operadora": empresa if empresa in {"VIVO", "TIM"} else "OUTRAS",
                "linha": texto(valores[colunas["LINHA"] - 1]),
                "chapaCpf": texto(valores[colunas["CHAPA/CPF"] - 1]),
                "nome": texto(valores[colunas["NOME"] - 1]),
                "cpf": texto(valores[colunas["CPF"] - 1]),
                "codCdc": texto(valores[colunas["COD CDC"] - 1]) or "SEM CODIGO",
                "cdc": texto(valores[colunas["CDC"] - 1]) or "SEM CENTRO DE CUSTO",
                "status": status,
                "valor": val_num,
            })
        return registros
    finally:
        wb.close()


def processar_historico(dados: list[dict[str, object]], gerado_em: datetime) -> dict[str, object]:
    """Salva a medição atual no histórico e retorna a comparação com a anterior."""
    HISTORICO_PATH.parent.mkdir(parents=True, exist_ok=True)
    
    totais_atuais = {
        "data": gerado_em.strftime("%d/%m/%Y às %H:%M"),
        "total_linhas": len(dados),
        "custo_total": sum(float(d["valor"]) for d in dados),
        "ATIVA": {"qtd": 0, "val": 0.0},
        "ESTOQUE": {"qtd": 0, "val": 0.0},
        "DESLIGADO": {"qtd": 0, "val": 0.0},
        "VERIFICAR": {"qtd": 0, "val": 0.0},
    }
    for d in dados:
        st = str(d["status"])
        if st in totais_atuais:
            totais_atuais[st]["qtd"] += 1
            totais_atuais[st]["val"] += float(d["valor"])

    historico = []
    if HISTORICO_PATH.exists():
        try:
            historico = json.loads(HISTORICO_PATH.read_text(encoding="utf-8"))
            if not isinstance(historico, list):
                historico = []
        except Exception:
            historico = []

    anterior = historico[-1] if historico else None
    
    def calc_delta(atual_val: float, ant_val: float) -> dict[str, object]:
        diff = atual_val - ant_val
        pct = (diff / ant_val * 100) if ant_val > 0 else 0.0
        return {"diff": diff, "pct": round(pct, 1), "tem_anterior": ant_val > 0}

    comparativo = {
        "tem_anterior": anterior is not None,
        "data_anterior": anterior.get("data") if anterior else None,
        "custo_total": calc_delta(totais_atuais["custo_total"], anterior.get("custo_total", 0.0) if anterior else 0.0),
        "ATIVA": calc_delta(totais_atuais["ATIVA"]["val"], anterior.get("ATIVA", {}).get("val", 0.0) if anterior else 0.0),
        "ESTOQUE": calc_delta(totais_atuais["ESTOQUE"]["val"], anterior.get("ESTOQUE", {}).get("val", 0.0) if anterior else 0.0),
        "DESLIGADO": calc_delta(totais_atuais["DESLIGADO"]["val"], anterior.get("DESLIGADO", {}).get("val", 0.0) if anterior else 0.0),
        "VERIFICAR": calc_delta(totais_atuais["VERIFICAR"]["val"], anterior.get("VERIFICAR", {}).get("val", 0.0) if anterior else 0.0),
    }

    # Evita gravar snapshots duplicados se rodar no mesmo minuto
    if not historico or historico[-1].get("data") != totais_atuais["data"]:
        historico.append(totais_atuais)
        if len(historico) > 30:  # Guarda os últimos 30 snapshots
            historico = historico[-30:]
        HISTORICO_PATH.write_text(json.dumps(historico, ensure_ascii=False, indent=2), encoding="utf-8")

    return comparativo


def exportar_dados(dados: list[dict[str, object]], comparativo: dict[str, object], gerado_em: datetime) -> None:
    """Exporta os dados em formato .js (para abrir direto com duplo clique) e .json."""
    SAIDAS_DIR.mkdir(parents=True, exist_ok=True)
    
    payload = {
        "gerado_em": gerado_em.strftime("%d/%m/%Y às %H:%M"),
        "total_registros": len(dados),
        "comparativo": comparativo,
        "dados": dados,
    }
    
    json_str = json.dumps(payload, ensure_ascii=False, indent=2)
    
    # 1. Salva arquivo .json
    DADOS_JSON_PADRAO.write_text(json_str, encoding="utf-8")
    
    # 2. Salva arquivo .js com variável global (permite duplo clique sem bloqueio CORS)
    js_content = f"window.DADOS_TELEFONIA = {json_str};\n"
    DADOS_JS_PADRAO.write_text(js_content, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--telefonia", type=Path, default=ARQUIVO_PADRAO)
    args = parser.parse_args()
    if not args.telefonia.exists():
        raise FileNotFoundError(f"Planilha de telefonia nao encontrada: {args.telefonia}")
    
    momento_atual = datetime.now()
    dados = carregar_dados(args.telefonia)
    comparativo = processar_historico(dados, momento_atual)
    exportar_dados(dados, comparativo, momento_atual)
    
    print("==================================================")
    print(f"Base de dados do Dashboard atualizada com sucesso!")
    print(f" • Arquivo JS:   {DADOS_JS_PADRAO}")
    print(f" • Arquivo JSON: {DADOS_JSON_PADRAO}")
    print(f" • Linhas carregadas: {len(dados)}")
    if comparativo["tem_anterior"]:
        print(f" • Comparativo com medição anterior ({comparativo['data_anterior']}):")
        print(f"   Variação Custo Total: {comparativo['custo_total']['pct']}%")
    print("==================================================")


if __name__ == "__main__":
    main()