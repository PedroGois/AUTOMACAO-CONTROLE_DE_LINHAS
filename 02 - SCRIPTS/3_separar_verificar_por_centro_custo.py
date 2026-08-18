import argparse
import copy
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


PASTA_SCRIPT = Path(__file__).resolve().parent
PASTA_PROJETO = PASTA_SCRIPT.parent
PASTA_DADOS = PASTA_PROJETO / "01 - DADOS"
SAIDA_PADRAO = PASTA_PROJETO / "04 - SAIDAS" / "VERIFICAR POR CENTRO DE CUSTO"
RESUMO_PADRAO = SAIDA_PADRAO / "resumo_separacao.txt"
def localizar_telefonia():
    """Usa a cópia de teste quando existir; na pasta oficial usa TELEFONIA.xlsx."""
    arquivo_teste = PASTA_DADOS / "TELEFONIA-TESTE.xlsx"
    return arquivo_teste if arquivo_teste.exists() else PASTA_DADOS / "TELEFONIA.xlsx"
TELEFONIA_PADRAO = localizar_telefonia()
def localizar_contatos():
    arquivo_teste = PASTA_DADOS / "CONTATO CDC-TESTE.xlsx"
    return arquivo_teste if arquivo_teste.exists() else PASTA_DADOS / "CONTATO CDC.xlsx"
CONTATOS_PADRAO = localizar_contatos()
ABA_ORIGEM = "Planos"
STATUS_ALVO = "VERIFICAR"
MARCADOR_RESUMO = "SEPARAÇÃO DOS VERIFICAR POR CENTRO DE CUSTO"

COR_CABECALHO = PatternFill("solid", fgColor="008C95")
FONTE_CABECALHO = Font(color="FFFFFF", bold=True)


def normalizar(valor):
    if valor is None:
        return ""
    texto = unicodedata.normalize("NFKD", str(valor).strip().upper())
    texto = "".join(letra for letra in texto if not unicodedata.combining(letra))
    texto = re.sub(r"[^A-Z0-9 ]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def nome_seguro(valor, limite=70):
    texto = normalizar(valor).replace(" ", "_")
    texto = re.sub(r"_+", "_", texto).strip("_")
    return (texto or "SEM_CENTRO_DE_CUSTO")[:limite]


def mapa_cabecalhos(ws):
    return {
        normalizar(ws.cell(1, coluna).value): coluna
        for coluna in range(1, ws.max_column + 1)
        if ws.cell(1, coluna).value is not None
    }


def carregar_responsaveis(caminho):
    """Retorna o responsável de cada Cod CDC a partir da planilha de contatos."""
    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        ws = wb.active
        cab = mapa_cabecalhos(ws)
        col_cod_cdc = cab.get("COD CDC")
        col_responsavel = cab.get("RESPONSAVEL")
        if not col_cod_cdc or not col_responsavel:
            raise ValueError("A planilha de contatos precisa das colunas Cod CDC e Responsavel.")
        responsaveis = {}
        for linha in range(2, ws.max_row + 1):
            cod_cdc = ws.cell(linha, col_cod_cdc).value
            responsavel = ws.cell(linha, col_responsavel).value
            if cod_cdc not in (None, ""):
                responsaveis[str(cod_cdc).strip()] = (
                    str(responsavel).strip()
                    if responsavel not in (None, "")
                    else "SEM_RESPONSAVEL"
                )
        return responsaveis
    finally:
        wb.close()


def copiar_estilo(origem, destino):
    if origem.has_style:
        destino._style = copy.copy(origem._style)
    destino.number_format = origem.number_format
    destino.alignment = copy.copy(origem.alignment)
    destino.font = copy.copy(origem.font)
    destino.fill = copy.copy(origem.fill)
    destino.border = copy.copy(origem.border)
    destino.protection = copy.copy(origem.protection)


def gerar_arquivo(ws_origem, linhas, caminho):
    wb_saida = Workbook()
    ws_saida = wb_saida.active
    ws_saida.title = "Verificar"
    ws_saida.freeze_panes = "A2"

    for coluna in range(1, ws_origem.max_column + 1):
        origem = ws_origem.cell(1, coluna)
        destino = ws_saida.cell(1, coluna, origem.value)
        copiar_estilo(origem, destino)
        destino.fill = copy.copy(COR_CABECALHO)
        destino.font = copy.copy(FONTE_CABECALHO)
        destino.alignment = Alignment(horizontal="center", vertical="center")

        letra = origem.column_letter
        largura = ws_origem.column_dimensions[letra].width
        ws_saida.column_dimensions[letra].width = min(max(largura or 12, 10), 45)

    for linha_saida, linha_origem in enumerate(linhas, start=2):
        for coluna in range(1, ws_origem.max_column + 1):
            origem = ws_origem.cell(linha_origem, coluna)
            destino = ws_saida.cell(linha_saida, coluna, origem.value)
            copiar_estilo(origem, destino)

    ws_saida.auto_filter.ref = ws_saida.dimensions
    tabela = Table(displayName="TabelaVerificar", ref=ws_saida.dimensions)
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws_saida.add_table(tabela)
    wb_saida.save(caminho)
    wb_saida.close()


def atualizar_resumo(caminho, arquivos, gerado_em, pasta_saida):
    conteudo = caminho.read_text(encoding="utf-8") if caminho.exists() else ""
    if MARCADOR_RESUMO in conteudo:
        conteudo = conteudo.split(MARCADOR_RESUMO, 1)[0].rstrip()

    linhas = [
        MARCADOR_RESUMO,
        f"Gerado em: {gerado_em:%d/%m/%Y às %H:%M:%S}",
        f"Arquivos por responsavel e Cod CDC: {len(arquivos)}",
        f"Linhas VERIFICAR separadas: {sum(qtd for _, qtd in arquivos)}",
        "",
        "Quantidade por arquivo:",
    ]
    linhas.extend(
        f"- {arquivo.relative_to(pasta_saida)}: {quantidade}"
        for arquivo, quantidade in arquivos
    )
    novo_conteudo = (conteudo + "\n\n" if conteudo else "") + "\n".join(linhas) + "\n"
    caminho.write_text(novo_conteudo, encoding="utf-8")


def separar(telefonia, contatos, pasta_saida, resumo):
    wb = load_workbook(telefonia, data_only=False)
    if ABA_ORIGEM not in wb.sheetnames:
        raise ValueError(f"A aba '{ABA_ORIGEM}' não foi encontrada em {telefonia}")

    ws = wb[ABA_ORIGEM]
    cab = mapa_cabecalhos(ws)
    col_status = cab.get("STATUS")
    col_cod_cdc = cab.get("COD CDC")
    col_cdc = cab.get("CDC")
    responsaveis = carregar_responsaveis(contatos)
    if not all((col_status, col_cod_cdc, col_cdc)):
        raise ValueError("As colunas Status, Cod CDC e CDC são obrigatórias.")

    grupos = defaultdict(list)
    nomes_cdc = defaultdict(list)
    for linha in range(2, ws.max_row + 1):
        if normalizar(ws.cell(linha, col_status).value) != STATUS_ALVO:
            continue
        cod_cdc = ws.cell(linha, col_cod_cdc).value
        cdc = ws.cell(linha, col_cdc).value
        chave = str(cod_cdc).strip() if cod_cdc not in (None, "") else "SEM_COD_CDC"
        grupos[chave].append(linha)
        nomes_cdc[chave].append(
            str(cdc).strip() if cdc not in (None, "") else "SEM_CENTRO_DE_CUSTO"
        )

    pasta_saida.mkdir(parents=True, exist_ok=True)
    for antigo in pasta_saida.rglob("VERIFICAR_*.xlsx"):
        antigo.unlink()

    arquivos = []
    for cod_cdc, linhas in sorted(grupos.items(), key=lambda item: item[0]):
        cdc = Counter(nomes_cdc[cod_cdc]).most_common(1)[0][0]
        responsavel = responsaveis.get(cod_cdc, "SEM_RESPONSAVEL")
        pasta_responsavel = pasta_saida / nome_seguro(responsavel)
        pasta_responsavel.mkdir(parents=True, exist_ok=True)
        nome = f"VERIFICAR_{nome_seguro(cod_cdc, 20)}_{nome_seguro(cdc)}.xlsx"
        caminho = pasta_responsavel / nome
        gerar_arquivo(ws, linhas, caminho)
        arquivos.append((caminho, len(linhas)))

    wb.close()
    gerado_em = datetime.now()
    atualizar_resumo(resumo, arquivos, gerado_em, pasta_saida)
    print(f"Pasta de saída: {pasta_saida.name}")
    print(f"Resumo atualizado: {resumo.name}")
    print(f"Arquivos gerados: {len(arquivos)}")
    print(f"Linhas VERIFICAR separadas: {sum(qtd for _, qtd in arquivos)}")
    for caminho, quantidade in arquivos:
        print(f"- {caminho.name}: {quantidade} linha(s)")


def main():
    parser = argparse.ArgumentParser(
        description="Separa as linhas VERIFICAR da aba Planos em arquivos XLSX por centro de custo."
    )
    parser.add_argument("--telefonia", type=Path, default=TELEFONIA_PADRAO)
    parser.add_argument("--contatos", type=Path, default=CONTATOS_PADRAO)
    parser.add_argument("--saida", type=Path, default=SAIDA_PADRAO)
    parser.add_argument("--resumo", type=Path, default=RESUMO_PADRAO)
    args = parser.parse_args()

    if not args.telefonia.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {args.telefonia}")
    if not args.contatos.exists():
        raise FileNotFoundError(f"Planilha de contatos nao encontrada: {args.contatos}")
    separar(args.telefonia, args.contatos, args.saida, args.resumo)


if __name__ == "__main__":
    main()
